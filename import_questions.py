from openpyxl import load_workbook

from app import create_app, db
from app.models import Category, Question

app = create_app()

with app.app_context():

    workbook = load_workbook("question2.xlsx")
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, values_only=True):

        category_name = str(row[0]).strip()
        difficulty = str(row[1]).strip()
        question_text = str(row[2]).strip()

        option1 = str(row[3]).strip()
        option2 = str(row[4]).strip()
        option3 = str(row[5]).strip()
        option4 = str(row[6]).strip()

        answer = str(row[7]).strip()

        category = Category.query.filter_by(
            category_name=category_name
        ).first()

        if category is None:

            category = Category(category_name=category_name)

            db.session.add(category)

            db.session.commit()

        question = Question(

            category_id=category.id,

            difficulty=difficulty,

            question=question_text,

            option1=option1,

            option2=option2,

            option3=option3,

            option4=option4,

            correct_answer=answer

        )

        db.session.add(question)

    db.session.commit()

print("✅ All Questions Imported Successfully!")